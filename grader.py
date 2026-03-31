import os
import re
from typing import List
import openpyxl
from models import GradingReport, SectionResult, CheckResult
from checkers.base_checker import check_base_structure
from checkers.sheet1 import check_sheet1
from checkers.sheet2 import check_sheet2
from checkers.sheet3 import check_sheet3
from checkers.sheet4 import check_sheet4

def grade_excel_file(file_path: str) -> GradingReport:
    filename = os.path.basename(file_path)
    
    # 1. 파일명 분석
    name_match = re.search(r'(\d+)-([가-힣]+)\.xlsx', filename)
    if name_match:
        student_id = name_match.group(1)
        student_name = name_match.group(2)
    else:
        student_id = "UnknownId"
        student_name = "UnknownName"

    sections: List[SectionResult] = []

    try:
        # data_only=False 설정으로 수식 열람
        wb = openpyxl.load_workbook(file_path, data_only=False)
    except Exception as e:
        print(f"Failed to load Excel file: {e}")
        return GradingReport(student_id, student_name, sections)
    
    # 공통 구조 검사 수행
    sections.append(SectionResult("기본 검사 (파일명/시트)", check_base_structure(file_path, wb)))
    
    # 제1작업
    if "제1작업" in wb.sheetnames:
        sections.append(SectionResult("제1작업", check_sheet1(wb["제1작업"], wb)))
    else:
        sections.append(SectionResult("제1작업", [CheckResult("제1작업 시트 존재 여부", False, 0, 10, "제1작업 시트가 없습니다.")]))

    # 제2작업
    if "제2작업" in wb.sheetnames:
        sections.append(SectionResult("제2작업", check_sheet2(wb["제2작업"], wb)))
    else:
        sections.append(SectionResult("제2작업", [CheckResult("제2작업 시트 존재 여부", False, 0, 10, "제2작업 시트가 없습니다.")]))
        
    # 제3작업
    if "제3작업" in wb.sheetnames:
        sections.append(SectionResult("제3작업", check_sheet3(wb["제3작업"], wb)))
    else:
        sections.append(SectionResult("제3작업", [CheckResult("제3작업 시트 존재 여부", False, 0, 10, "제3작업 시트가 없습니다.")]))
        
    # 제4작업
    if "제4작업" in wb.sheetnames:
        sections.append(SectionResult("제4작업", check_sheet4(wb["제4작업"], wb)))
    else:
        sections.append(SectionResult("제4작업", [CheckResult("제4작업 시트 존재 여부", False, 0, 10, "제4작업 시트가 없습니다.")]))

    wb.close()
    return GradingReport(student_id, student_name, sections)
